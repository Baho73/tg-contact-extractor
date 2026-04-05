"""Tests for src.to_excel — JSON-to-Excel conversion."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from src.to_excel import json_to_excel, normalize_type


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

SAMPLE_DATA = {
    "extraction_meta": {
        "source": "test.json",
        "total_messages": 100,
        "processed_batches": 2,
        "skipped_batches": [],
        "date": "2026-04-06",
        "prompt": "",
    },
    "contacts": [
        {
            "source_message_id": 1,
            "from_user": "Alice",
            "context": "job post",
            "extracted": [
                {"type": "вакансия", "value": "Python dev"},
                {"type": "email", "value": "hr@test.ru"},
                {"type": "компания", "value": "TestCo"},
            ],
        },
        {
            "source_message_id": 2,
            "from_user": "Bob",
            "context": "recruiter",
            "extracted": [
                {"type": "рекрутер", "value": "Иванов"},
                {"type": "telegram", "value": "@ivan"},
            ],
        },
    ],
}


def _write_json(tmp_path: Path, data: dict | None = None) -> Path:
    """Write sample JSON and return the path."""
    json_path = tmp_path / "contacts.json"
    json_path.write_text(json.dumps(data or SAMPLE_DATA, ensure_ascii=False), encoding="utf-8")
    return json_path


# ---------------------------------------------------------------------------
# normalize_type
# ---------------------------------------------------------------------------


def test_normalize_type():
    assert normalize_type("телеграм") == "telegram"
    assert normalize_type("емаил") == "email"
    assert normalize_type("WhatsApp") == "whatsapp"
    # unknown type passes through unchanged
    assert normalize_type("неизвестный") == "неизвестный"


# ---------------------------------------------------------------------------
# json_to_excel — file creation
# ---------------------------------------------------------------------------


def test_json_to_excel_creates_file(tmp_path: Path):
    json_path = _write_json(tmp_path)
    xlsx_path = json_to_excel(json_path)
    assert xlsx_path.exists()
    assert xlsx_path.suffix == ".xlsx"


# ---------------------------------------------------------------------------
# json_to_excel — column headers
# ---------------------------------------------------------------------------


def test_json_to_excel_columns(tmp_path: Path):
    json_path = _write_json(tmp_path)
    xlsx_path = json_to_excel(json_path)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Контакты"]
    headers = [cell.value for cell in ws[1]]

    # Expected data columns based on COLUMN_ORDER + metadata columns
    assert "Вакансия" in headers
    assert "Компания" in headers
    assert "Email" in headers
    assert "Рекрутер" in headers
    assert "Telegram" in headers
    assert "Контекст" in headers
    assert "Id Сообщения" in headers
    wb.close()


# ---------------------------------------------------------------------------
# json_to_excel — row count
# ---------------------------------------------------------------------------


def test_json_to_excel_row_count(tmp_path: Path):
    json_path = _write_json(tmp_path)
    xlsx_path = json_to_excel(json_path)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Контакты"]
    # row 1 = header, rows 2-3 = 2 contacts
    data_rows = ws.max_row - 1
    assert data_rows == 2
    wb.close()


# ---------------------------------------------------------------------------
# json_to_excel — summary sheet exists
# ---------------------------------------------------------------------------


def test_json_to_excel_summary_sheet(tmp_path: Path):
    json_path = _write_json(tmp_path)
    xlsx_path = json_to_excel(json_path)

    wb = openpyxl.load_workbook(xlsx_path)
    assert "Сводка" in wb.sheetnames

    ws = wb["Сводка"]

    # Check that source metadata appears in the summary
    values_col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "Источник" in values_col_a
    assert "Дата обработки" in values_col_a
    assert "Извлечено контактов" in values_col_a
    wb.close()


# ---------------------------------------------------------------------------
# json_to_excel — prompt in summary
# ---------------------------------------------------------------------------


def test_json_to_excel_prompt_in_summary(tmp_path: Path):
    data = json.loads(json.dumps(SAMPLE_DATA))
    data["extraction_meta"]["prompt"] = "test prompt"
    json_path = _write_json(tmp_path, data)
    xlsx_path = json_to_excel(json_path)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Сводка"]

    # Find "Промпт" label
    found = False
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Промпт":
            assert ws.cell(row=row, column=2).value == "test prompt"
            found = True
            break
    assert found, "Промпт row not found in summary sheet"
    wb.close()


# ---------------------------------------------------------------------------
# json_to_excel — empty contacts
# ---------------------------------------------------------------------------


def test_json_to_excel_empty_contacts(tmp_path: Path):
    data = {
        "extraction_meta": {"source": "empty.json", "total_messages": 0},
        "contacts": [],
    }
    json_path = _write_json(tmp_path, data)
    xlsx_path = json_to_excel(json_path)
    assert xlsx_path.exists()

    wb = openpyxl.load_workbook(xlsx_path)
    assert "Контакты" in wb.sheetnames
    assert "Сводка" in wb.sheetnames
    wb.close()
