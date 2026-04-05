# FILE: src/to_excel.py
# VERSION: 1.0.0
# START_MODULE_CONTRACT
#   PURPOSE: Convert extracted contacts JSON to a formatted Excel file
#   SCOPE: JSON reading, type normalization, Excel generation
#   DEPENDS: M-WRITER (JSON format)
#   LINKS: M-EXCEL
# END_MODULE_CONTRACT
#
# START_MODULE_MAP
#   normalize_type — normalize contact type strings
#   json_to_excel — convert JSON file to Excel
#   main — CLI entry point
# END_MODULE_MAP

from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# START_BLOCK_TYPE_NORMALIZATION
TYPE_ALIASES: dict[str, str] = {
    "телеграм": "telegram",
    "Telegram": "telegram",
    "емаил": "email",
    "Email": "email",
    "Телефон": "телефон",
    "whatsApp": "whatsapp",
    "WhatsApp": "whatsapp",
    "LinkedIn": "linkedin",
    "Linkedin": "linkedin",
    "скайп": "skype",
    "Skype": "skype",
}

# Column order for the Excel file — priority fields first
COLUMN_ORDER: list[str] = [
    "вакансия",
    "компания",
    "город",
    "зарплата",
    "рекрутер",
    "email",
    "telegram",
    "телефон",
    "whatsapp",
    "linkedin",
    "сайт",
]

COLUMN_HEADERS: dict[str, str] = {
    "вакансия": "Вакансия",
    "компания": "Компания",
    "город": "Город",
    "зарплата": "Зарплата",
    "рекрутер": "Рекрутер",
    "email": "Email",
    "telegram": "Telegram",
    "телефон": "Телефон",
    "whatsapp": "WhatsApp",
    "linkedin": "LinkedIn",
    "сайт": "Сайт",
}
# END_BLOCK_TYPE_NORMALIZATION


# START_CONTRACT: normalize_type
#   PURPOSE: Normalize contact type string via alias map
#   INPUTS: { raw_type: str }
#   OUTPUTS: str
# END_CONTRACT: normalize_type
def normalize_type(raw_type: str) -> str:
    return TYPE_ALIASES.get(raw_type, raw_type)


# START_BLOCK_STYLE_CONSTANTS
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Calibri", size=10)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
STRIPE_FILL = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid")
# END_BLOCK_STYLE_CONSTANTS


# START_CONTRACT: json_to_excel
#   PURPOSE: Convert contacts JSON to formatted Excel workbook
#   INPUTS: { json_path: Path, output_path: Path | None }
#   OUTPUTS: { Path — path to written .xlsx file }
#   SIDE_EFFECTS: Writes .xlsx to disk
#   LINKS: M-EXCEL
# END_CONTRACT: json_to_excel
def json_to_excel(json_path: Path, output_path: Path | None = None) -> Path:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    contacts = data.get("contacts", [])
    meta = data.get("extraction_meta", {})

    # START_BLOCK_DISCOVER_COLUMNS
    # Discover all types present in data
    all_types: set[str] = set()
    for c in contacts:
        for f in c.get("extracted", []):
            all_types.add(normalize_type(f["type"]))

    # Build column list: known order first, then any extras alphabetically
    columns: list[str] = [col for col in COLUMN_ORDER if col in all_types]
    extras = sorted(all_types - set(COLUMN_ORDER))
    columns.extend(extras)
    # END_BLOCK_DISCOVER_COLUMNS

    # START_BLOCK_BUILD_ROWS
    rows: list[dict[str, str]] = []
    for c in contacts:
        row: dict[str, str] = {}
        for f in c.get("extracted", []):
            t = normalize_type(f["type"])
            value = f["value"]
            # If same type appears multiple times, join with separator
            if t in row:
                row[t] = row[t] + " | " + value
            else:
                row[t] = value
        # Add context as a hidden helper
        row["_context"] = c.get("context", "")
        row["_source_id"] = str(c.get("source_message_id", ""))
        row["_from_user"] = c.get("from_user", "")
        rows.append(row)
    # END_BLOCK_BUILD_ROWS

    # Add metadata columns at the end
    all_columns = columns + ["Контекст", "ID сообщения"]

    # START_BLOCK_WRITE_EXCEL
    wb = Workbook()
    ws = wb.active
    ws.title = "Контакты"

    # -- Header row --
    for col_idx, col_key in enumerate(all_columns, 1):
        header_text = COLUMN_HEADERS.get(col_key, col_key.title())
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # -- Data rows --
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_key in enumerate(all_columns, 1):
            if col_key == "Контекст":
                value = row_data.get("_context", "")
            elif col_key == "ID сообщения":
                value = row_data.get("_source_id", "")
            else:
                value = row_data.get(col_key, "")

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

            # Zebra striping
            if row_idx % 2 == 0:
                cell.fill = STRIPE_FILL

    # -- Auto-width columns --
    for col_idx in range(1, len(all_columns) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, min(len(rows) + 2, 100)):  # sample first 100 rows
            cell = ws[f"{col_letter}{row_idx}"]
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, 10), 50)
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # -- Auto-filter --
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_columns))}{len(rows) + 1}"

    # -- Summary sheet --
    ws_summary = wb.create_sheet("Сводка")
    ws_summary.cell(row=1, column=1, value="Параметр").font = Font(bold=True)
    ws_summary.cell(row=1, column=2, value="Значение").font = Font(bold=True)

    summary_data = [
        ("Источник", meta.get("source", "")),
        ("Дата обработки", meta.get("date", "")),
        ("Всего сообщений", meta.get("total_messages", 0)),
        ("Обработано пакетов", meta.get("processed_batches", 0)),
        ("Пропущено пакетов", len(meta.get("skipped_batches", []))),
        ("Извлечено контактов", len(rows)),
        ("", ""),
        ("Типы данных", "Количество"),
    ]

    type_counts: dict[str, int] = {}
    for r in rows:
        for col in columns:
            if r.get(col):
                type_counts[col] = type_counts.get(col, 0) + 1

    for t in columns:
        summary_data.append((COLUMN_HEADERS.get(t, t), type_counts.get(t, 0)))

    # Add prompt used
    prompt_text = meta.get("prompt", "")
    if prompt_text:
        summary_data.append(("", ""))
        summary_data.append(("Промпт", prompt_text))

    for i, (param, val) in enumerate(summary_data, 2):
        ws_summary.cell(row=i, column=1, value=param)
        ws_summary.cell(row=i, column=2, value=val)

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 50
    # END_BLOCK_WRITE_EXCEL

    # Save
    if output_path is None:
        output_path = json_path.with_suffix(".xlsx")

    wb.save(output_path)
    logger.info("[Excel][json_to_excel] Wrote %d rows to %s", len(rows), output_path)
    return output_path


# START_CONTRACT: main
#   PURPOSE: CLI entry point for JSON-to-Excel conversion
#   INPUTS: sys.argv
#   OUTPUTS: None
#   SIDE_EFFECTS: writes .xlsx
# END_CONTRACT: main
def main() -> None:
    parser = argparse.ArgumentParser(description="Convert contacts JSON to Excel")
    parser.add_argument("json_path", help="Path to contacts_*.json")
    parser.add_argument("-o", "--output", help="Output .xlsx path (default: same name)")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO)

    json_path = Path(args.json_path)
    output = Path(args.output) if args.output else None
    result = json_to_excel(json_path, output)
    print(f"Done! {result}")


if __name__ == "__main__":
    main()


# START_CHANGE_SUMMARY
#   LAST_CHANGE: [v1.0.0 — Initial: JSON->Excel with type normalization, auto-columns, zebra, summary sheet]
# END_CHANGE_SUMMARY
